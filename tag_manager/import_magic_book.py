from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .db import PROJECT_DIR, clean_text, init_db, upsert_group, upsert_tag

MAGIC_BOOK_PATH = PROJECT_DIR / "提示词" / "魔导书.xlsx"
TAG_PATTERN = re.compile(r"^[A-Za-z0-9_!?'’().,\- /:{}\[\]\\&]+$")
CATEGORY_BY_SHEET = {
    "1.镜头": "1.镜头",
    "2.人物": "2.人物",
    "3.服饰": "3.服饰",
    "4.表情": "4.表情",
    "5.动作": "5.动作",
    "6.场景道具": "6.场景道具",
    "0.可以涩涩": "0.可以涩涩",
    "NSFW Tags": "NSFW Tags",
}


def looks_like_tag(value: str) -> bool:
    value = clean_text(value)
    if not value or len(value) > 120:
        return False
    if any("一" <= ch <= "鿿" for ch in value):
        return False
    if not TAG_PATTERN.match(value):
        return False
    return any(ch.isalpha() for ch in value)


def is_subcategory_header(value: str) -> bool:
    value = clean_text(value)
    if not value or looks_like_tag(value):
        return False
    if len(value) > 24:
        return False
    if any(mark in value for mark in ("。", "，", "、", "；", "：", ":", "（", "）", "(", ")", "<", ">")):
        return False
    return True


def import_sheet_pairs(ws, category: str) -> int:
    count = 0
    source = f"魔导书.xlsx:{ws.title}"
    current_headers: dict[int, str] = {}
    for row in ws.iter_rows(values_only=True):
        values = [clean_text(v) for v in row]
        for idx in range(0, len(values) - 1, 2):
            first = values[idx]
            second = values[idx + 1]
            if not first and not second:
                continue
            if first and not looks_like_tag(first):
                if is_subcategory_header(first):
                    current_headers[idx] = first
                continue
            if looks_like_tag(first):
                subcategory = current_headers.get(idx, "未分类") or "未分类"
                upsert_tag(tag=first, zh=second, category=category, subcategory=subcategory, source=source)
                count += 1
    return count


def import_tag_groups(ws) -> int:
    count = 0
    source = f"魔导书.xlsx:{ws.title}"
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        values = [clean_text(v) for v in row]
        if len(values) < 3:
            continue
        category, name, positive = values[0], values[1], values[2]
        negative = values[3] if len(values) > 3 else ""
        notes = values[4] if len(values) > 4 else ""
        if name and positive:
            upsert_group(name=name, category=category, positive_prompt=positive, negative_prompt=negative, notes=notes, source=source)
            count += 1
    return count


def import_magic_book(path: Path = MAGIC_BOOK_PATH) -> dict[str, int]:
    init_db()
    wb = load_workbook(path, data_only=True, read_only=True)
    result: dict[str, int] = {}
    for ws in wb.worksheets:
        if ws.title == "tag组合":
            result[ws.title] = import_tag_groups(ws)
        elif ws.title in CATEGORY_BY_SHEET:
            result[ws.title] = import_sheet_pairs(ws, CATEGORY_BY_SHEET[ws.title])
    return result


if __name__ == "__main__":
    imported = import_magic_book()
    for sheet, count in imported.items():
        print(f"{sheet}: {count}")
